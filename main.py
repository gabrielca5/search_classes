"""
Sistema de Monitoramento de Salas - Insper
Autor: Sistema Automatizado
Data: 2025-10-09
Descrição: Busca e registra todas as salas (disponíveis e ocupadas) por horário
"""

import json
import logging
import re
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Set
from dataclasses import dataclass, asdict

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================================
# CONFIGURAÇÕES
# ============================================================================
XML_URL = "https://cgi.insper.edu.br/Agenda/xml/ExibeCalendario.xml"
MAX_RETRIES = 3
REQUEST_TIMEOUT = 15
OUTPUT_DIR = Path("output")

# Configuração de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# ============================================================================
# MODELOS DE DADOS
# ============================================================================
@dataclass
class Sala:
    """Modelo de dados para sala"""
    data: str
    dia_semana: str
    horario: str
    sala: str
    predio: str
    curso: Optional[str] = None  # Apenas para salas ocupadas

    def to_dict(self) -> Dict:
        """Converte para dicionário"""
        return asdict(self)


# ============================================================================
# FUNÇÕES AUXILIARES
# ============================================================================
def obter_dia_semana(data: datetime) -> str:
    """
    Retorna o nome do dia da semana em português
    
    Args:
        data: Objeto datetime
        
    Returns:
        Nome do dia da semana
    """
    dias = {
        0: "Segunda-feira",
        1: "Terça-feira",
        2: "Quarta-feira",
        3: "Quinta-feira",
        4: "Sexta-feira",
        5: "Sábado",
        6: "Domingo"
    }
    return dias[data.weekday()]


def gerar_timestamp() -> str:
    """
    Gera timestamp para nome de arquivo
    
    Returns:
        String no formato YYYY-MM-DD_HH-MM-SS
    """
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")


def eh_sala_valida(sala: str) -> bool:
    """
    Verifica se a sala é válida (qualquer formato)
    
    Args:
        sala: String da sala
        
    Returns:
        True se a sala for válida
    """
    return bool(sala and sala.strip())


def converter_horario_para_minutos(horario_str: str) -> int:
    """
    Converte string de horário para minutos desde meia-noite
    
    Args:
        horario_str: String no formato "HH:MM"
        
    Returns:
        Minutos desde meia-noite
    """
    try:
        h, m = map(int, horario_str.split(':'))
        return h * 60 + m
    except:
        return 0


def agrupar_horarios_30min(todos_horarios: Set[str]) -> List[str]:
    """
    Agrupa horários em blocos de 30 minutos
    
    Args:
        todos_horarios: Conjunto de todos os horários
        
    Returns:
        Lista de horários agrupados
    """
    horarios_inicio = set()
    for horario in todos_horarios:
        inicio = horario.split(' - ')[0]
        horarios_inicio.add(inicio)
    
    horarios_agrupados = set()
    for horario in sorted(horarios_inicio):
        minutos = converter_horario_para_minutos(horario)
        bloco_minutos = (minutos // 30) * 30
        h = bloco_minutos // 60
        m = bloco_minutos % 60
        horarios_agrupados.add(f"{h:02d}:{m:02d}")
    
    return sorted(list(horarios_agrupados))


# ============================================================================
# FUNÇÕES DE SCRAPING
# ============================================================================
def buscar_xml_com_retry(url: str, max_tentativas: int = MAX_RETRIES) -> Optional[bytes]:
    """
    Busca o conteúdo XML com retry
    
    Args:
        url: URL do XML
        max_tentativas: Número máximo de tentativas
        
    Returns:
        Conteúdo do XML ou None em caso de falha
    """
    for tentativa in range(1, max_tentativas + 1):
        try:
            logger.info(f"Tentativa {tentativa}/{max_tentativas}: Buscando dados do XML...")
            response = requests.get(url, timeout=REQUEST_TIMEOUT)
            response.raise_for_status()
            logger.info("✓ Dados obtidos com sucesso!")
            return response.content
        except requests.exceptions.RequestException as e:
            logger.error(f"✗ Erro na tentativa {tentativa}: {e}")
            if tentativa == max_tentativas:
                logger.error("✗ Número máximo de tentativas atingido. Falha ao obter dados.")
                return None
    return None


def extrair_dados_xml(xml_content: bytes) -> tuple:
    """
    Extrai dados do XML
    
    Args:
        xml_content: Conteúdo do XML em bytes
        
    Returns:
        Tupla com (mapa_ocupacao_completo, mapa_predios, horarios_agrupados)
    """
    logger.info("Analisando dados do XML...")
    soup = BeautifulSoup(xml_content, "lxml-xml")
    
    # Estrutura: horario -> sala -> {predio, curso}
    mapa_ocupacao_completo: Dict[str, Dict[str, Dict[str, str]]] = {}
    mapa_predios: Dict[str, str] = {}
    todos_horarios_originais: Set[str] = set()
    
    eventos = soup.find_all("CalendarioEvento")
    logger.info(f"Total de eventos encontrados: {len(eventos)}")
    
    for evento in eventos:
        try:
            sala = evento.find("sala").text.strip()
            if not eh_sala_valida(sala):
                continue
            
            predio = evento.find("predio").text.strip()
            hora_inicio = evento.find("horainicio").text.strip()
            hora_termino = evento.find("horatermino").text.strip()
            curso = evento.find("turma").text.strip()
            
            horario_completo = f"{hora_inicio} - {hora_termino}"
            
            # Mapear prédio da sala
            mapa_predios[sala] = predio
            
            # Mapear ocupação completa
            if horario_completo not in mapa_ocupacao_completo:
                mapa_ocupacao_completo[horario_completo] = {}
            
            mapa_ocupacao_completo[horario_completo][sala] = {
                "predio": predio,
                "curso": curso
            }
            
            todos_horarios_originais.add(horario_completo)
            
        except (AttributeError, KeyError) as e:
            logger.warning(f"Erro ao processar evento: {e}")
            continue
    
    # Agrupar horários em blocos de 30 minutos
    horarios_agrupados = agrupar_horarios_30min(todos_horarios_originais)
    
    logger.info(f"Total de salas mapeadas: {len(mapa_predios)}")
    logger.info(f"Total de horários agrupados (30 min): {len(horarios_agrupados)}")
    
    return mapa_ocupacao_completo, mapa_predios, horarios_agrupados


def processar_disponibilidade(
    mapa_ocupacao_completo: Dict[str, Dict[str, Dict[str, str]]],
    mapa_predios: Dict[str, str],
    horarios_agrupados: List[str],
    data_ref: datetime
) -> Dict[str, List[Sala]]:
    """
    Processa a disponibilidade de salas
    
    Args:
        mapa_ocupacao_completo: Mapa completo de ocupação
        mapa_predios: Mapa de prédios por sala
        horarios_agrupados: Lista de horários agrupados
        data_ref: Data de referência
        
    Returns:
        Dicionário com salas_disponiveis e salas_ocupadas
    """
    logger.info("Processando disponibilidade de salas...")
    
    salas_disponiveis: List[Sala] = []
    salas_ocupadas: List[Sala] = []
    
    data_str = data_ref.strftime("%Y-%m-%d")
    dia_semana = obter_dia_semana(data_ref)
    
    todas_salas_sistema = set(mapa_predios.keys())
    
    # Criar um mapa de ocupação por bloco de 30 minutos
    for horario_bloco in horarios_agrupados:
        minuto_bloco = converter_horario_para_minutos(horario_bloco)
        minuto_fim = minuto_bloco + 30
        h_fim = minuto_fim // 60
        m_fim = minuto_fim % 60
        horario_formatado = f"{horario_bloco} - {h_fim:02d}:{m_fim:02d}"
        
        # Encontrar todas as aulas que estão acontecendo neste bloco
        salas_ocupadas_neste_bloco = {}
        
        for horario_aula, salas_dict in mapa_ocupacao_completo.items():
            inicio_aula_str = horario_aula.split(' - ')[0]
            fim_aula_str = horario_aula.split(' - ')[1]
            
            inicio_aula = converter_horario_para_minutos(inicio_aula_str)
            fim_aula = converter_horario_para_minutos(fim_aula_str)
            
            # Verificar se a aula está acontecendo durante este bloco
            if inicio_aula < minuto_fim and fim_aula > minuto_bloco:
                for sala, info in salas_dict.items():
                    salas_ocupadas_neste_bloco[sala] = info
        
        # Separar salas ocupadas e disponíveis
        for sala in sorted(todas_salas_sistema):
            predio = mapa_predios.get(sala, "Prédio não identificado")
            
            if sala in salas_ocupadas_neste_bloco:
                # Sala ocupada
                info = salas_ocupadas_neste_bloco[sala]
                salas_ocupadas.append(
                    Sala(
                        data=data_str,
                        dia_semana=dia_semana,
                        horario=horario_formatado,
                        sala=sala,
                        predio=info["predio"],
                        curso=info["curso"]
                    )
                )
            else:
                # Sala disponível
                salas_disponiveis.append(
                    Sala(
                        data=data_str,
                        dia_semana=dia_semana,
                        horario=horario_formatado,
                        sala=sala,
                        predio=predio
                    )
                )
    
    logger.info(f"Salas disponíveis: {len(salas_disponiveis)}")
    logger.info(f"Salas ocupadas: {len(salas_ocupadas)}")
    
    return {
        "salas_disponiveis": salas_disponiveis,
        "salas_ocupadas": salas_ocupadas
    }


# ============================================================================
# FUNÇÕES DE EXPORTAÇÃO
# ============================================================================
def salvar_json(dados: Dict[str, List[Sala]], timestamp: str) -> Path:
    """
    Salva dados em arquivo JSON
    
    Args:
        dados: Dicionário com salas disponíveis e ocupadas
        timestamp: Timestamp para nome do arquivo
        
    Returns:
        Caminho do arquivo salvo
    """
    OUTPUT_DIR.mkdir(exist_ok=True)
    arquivo = OUTPUT_DIR / f"salas_disponiveis_{timestamp}.json"
    
    dados_dict = {
        "salas_disponiveis": [s.to_dict() for s in dados["salas_disponiveis"]],
        "salas_ocupadas": [s.to_dict() for s in dados["salas_ocupadas"]]
    }
    
    with open(arquivo, 'w', encoding='utf-8') as f:
        json.dump(dados_dict, f, ensure_ascii=False, indent=2)
    
    logger.info(f"✓ JSON salvo: {arquivo}")
    return arquivo


def criar_xlsx_estilizado(dados: Dict[str, List[Sala]], timestamp: str) -> Path:
    """
    Cria arquivo XLSX com formatação visual moderna
    
    Args:
        dados: Dicionário com salas disponíveis e ocupadas
        timestamp: Timestamp para nome do arquivo
        
    Returns:
        Caminho do arquivo salvo
    """
    OUTPUT_DIR.mkdir(exist_ok=True)
    arquivo = OUTPUT_DIR / f"salas_disponiveis_{timestamp}.xlsx"
    
    wb = openpyxl.Workbook()
    
    # Cores modernas e profissionais
    cor_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    cor_header_ocupadas = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
    cor_livre = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    cor_livre_alt = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    cor_ocupada = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    cor_ocupada_alt = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    
    font_header = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    font_normal = Font(name='Calibri', size=11)
    
    alignment_center = Alignment(horizontal='center', vertical='center')
    alignment_left = Alignment(horizontal='left', vertical='center')
    
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # ====== ABA 1: SALAS DISPONÍVEIS ======
    ws_disponiveis = wb.active
    ws_disponiveis.title = "Salas Disponíveis"
    
    headers_disponiveis = ["Data", "Dia da Semana", "Horário", "Sala", "Prédio"]
    ws_disponiveis.append(headers_disponiveis)
    
    for col_num, header in enumerate(headers_disponiveis, 1):
        cell = ws_disponiveis.cell(row=1, column=col_num)
        cell.fill = cor_header
        cell.font = font_header
        cell.alignment = alignment_center
        cell.border = border
    
    for idx, sala in enumerate(dados["salas_disponiveis"], start=2):
        data_formatada = datetime.strptime(sala.data, "%Y-%m-%d").strftime("%d/%m/%Y")
        row = [data_formatada, sala.dia_semana, sala.horario, sala.sala, sala.predio]
        ws_disponiveis.append(row)
        
        for col_num in range(1, 6):
            cell = ws_disponiveis.cell(row=idx, column=col_num)
            cell.border = border
            cell.font = font_normal
            cell.fill = cor_livre if idx % 2 == 0 else cor_livre_alt
            cell.alignment = alignment_center if col_num in [1, 3] else alignment_left
    
    # Ajustar largura
    ws_disponiveis.column_dimensions['A'].width = 12
    ws_disponiveis.column_dimensions['B'].width = 18
    ws_disponiveis.column_dimensions['C'].width = 18
    ws_disponiveis.column_dimensions['D'].width = 10
    ws_disponiveis.column_dimensions['E'].width = 30
    ws_disponiveis.freeze_panes = 'A2'
    
    # ====== ABA 2: SALAS OCUPADAS ======
    ws_ocupadas = wb.create_sheet("Salas Ocupadas")
    
    headers_ocupadas = ["Data", "Dia da Semana", "Horário", "Sala", "Prédio", "Curso"]
    ws_ocupadas.append(headers_ocupadas)
    
    for col_num, header in enumerate(headers_ocupadas, 1):
        cell = ws_ocupadas.cell(row=1, column=col_num)
        cell.fill = cor_header_ocupadas
        cell.font = font_header
        cell.alignment = alignment_center
        cell.border = border
    
    for idx, sala in enumerate(dados["salas_ocupadas"], start=2):
        data_formatada = datetime.strptime(sala.data, "%Y-%m-%d").strftime("%d/%m/%Y")
        row = [data_formatada, sala.dia_semana, sala.horario, sala.sala, sala.predio, sala.curso]
        ws_ocupadas.append(row)
        
        for col_num in range(1, 7):
            cell = ws_ocupadas.cell(row=idx, column=col_num)
            cell.border = border
            cell.font = font_normal
            cell.fill = cor_ocupada if idx % 2 == 0 else cor_ocupada_alt
            cell.alignment = alignment_center if col_num in [1, 3] else alignment_left
    
    # Ajustar largura
    ws_ocupadas.column_dimensions['A'].width = 12
    ws_ocupadas.column_dimensions['B'].width = 18
    ws_ocupadas.column_dimensions['C'].width = 18
    ws_ocupadas.column_dimensions['D'].width = 10
    ws_ocupadas.column_dimensions['E'].width = 30
    ws_ocupadas.column_dimensions['F'].width = 35
    ws_ocupadas.freeze_panes = 'A2'
    
    wb.save(arquivo)
    logger.info(f"✓ XLSX salvo: {arquivo}")
    return arquivo


# ============================================================================
# FUNÇÃO PRINCIPAL
# ============================================================================
def main():
    """Função principal do sistema"""
    logger.info("=" * 70)
    logger.info("SISTEMA DE MONITORAMENTO DE SALAS - INSPER")
    logger.info("=" * 70)
    
    # 1. Buscar XML
    xml_content = buscar_xml_com_retry(XML_URL)
    if not xml_content:
        logger.error("Não foi possível obter os dados. Encerrando.")
        return
    
    # 2. Extrair dados
    mapa_ocupacao_completo, mapa_predios, horarios_agrupados = extrair_dados_xml(xml_content)
    
    if not horarios_agrupados:
        logger.warning("Nenhum horário encontrado no XML.")
        return
    
    # 3. Processar disponibilidade
    data_hoje = datetime.now()
    resultado = processar_disponibilidade(
        mapa_ocupacao_completo, mapa_predios, horarios_agrupados, data_hoje
    )
    
    if not resultado["salas_disponiveis"] and not resultado["salas_ocupadas"]:
        logger.warning("Nenhuma sala encontrada.")
        return
    
    # 4. Gerar timestamp
    timestamp = gerar_timestamp()
    
    # 5. Exportar dados
    logger.info("\nExportando dados...")
    arquivo_json = salvar_json(resultado, timestamp)
    arquivo_xlsx = criar_xlsx_estilizado(resultado, timestamp)
    
    # 6. Resumo final
    logger.info("\n" + "=" * 70)
    logger.info("RESUMO DA EXECUÇÃO")
    logger.info("=" * 70)
    logger.info(f"Data processada: {data_hoje.strftime('%d/%m/%Y')} ({obter_dia_semana(data_hoje)})")
    logger.info(f"Total de salas disponíveis: {len(resultado['salas_disponiveis'])}")
    logger.info(f"Total de salas ocupadas: {len(resultado['salas_ocupadas'])}")
    logger.info(f"Arquivos gerados:")
    logger.info(f"  • JSON: {arquivo_json}")
    logger.info(f"  • XLSX: {arquivo_xlsx}")
    logger.info("=" * 70)
    logger.info("✓ Processo concluído com sucesso!")


if __name__ == "__main__":
    main()